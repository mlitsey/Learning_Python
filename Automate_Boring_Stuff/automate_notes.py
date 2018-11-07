
"Chapter 4"
"Augmented assignment statement Equivalent assignment statement"
spam = spam + 1 : spam += 1
spam = spam - 1 : spam -= 1
spam = spam * 1 : spam *= 1
spam = spam / 1 : spam /= 1
spam = spam % 1 : spam %= 1

list = []
tuple = (0,1)

"Chapter 5"
dictionary = {}
keys(), values(), and items() methods

"Chapter 6"
Escape character Prints 'as'
'\' : Single quote'
'\" : Double quote'
'\t : Tab'
'\n : Newline (line break)'
'\\ : Backslash'

raw string = print(r'That is Carol\'s cat.')

"Chapter 12"
from openpyxl.utils import get_column_letter, column_index_from_string


for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')




import openpyxl
wb = openpyxl.Workbook()
wb.sheetnames
    ['Sheet']
sheet = wb.active
sheet.title
   'Sheet'
sheet.title = 'Spam Bacon Eggs Sheet'
wb.sheetnames
   ['Spam Bacon Eggs Sheet']



import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx')




import openpyxl
wb = openpyxl.Workbook()
wb.sheetnames
wb.create_sheet()
wb.sheetnames
wb.create_sheet(index=0, title='First Sheet')
wb.sheetnames
wb.create_sheet(index=2, title='Middle Sheet')
wb.sheetnames

wb.remove(wb['Middle Sheet'])   #### wb.remove_sheet(wb.sheetnames('Middle Sheet'))
wb.remove(wb['Sheet1']) or del wb['Sheet1'] #### wb.remove_sheet(wb.sheetnames('Sheet1'))
wb.sheetnames

